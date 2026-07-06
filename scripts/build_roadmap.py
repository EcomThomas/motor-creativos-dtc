# -*- coding: utf-8 -*-
"""
[LEGACY / DEPRECADO] — El motor ya NO vuelca al Excel por defecto (se hacía a mano y el
archivo se corrompía / era demasiada info). El entregable de Fase 3 es ahora el formato
ClickUp (tarea madre + subtareas) de `clickup_export.py`. Este script se conserva solo por
si alguien quiere poblar un Excel manualmente; no forma parte del flujo estándar.

build_roadmap.py — Vuelca baches a la hoja "Creative Roadmap" de un Excel, respetando
los valores de dropdown (validaciones) del template. Product-agnostic.

Uso:
    python build_roadmap.py <xlsx> <batches.json> [sheet="Creative Roadmap"] [start_row=auto] [author="AI"]

<batches.json> = lista de baches:
[
  {
    "concept": "...", "angle": "...", "avatar": "...", "mass_desire": "Quiero...",
    "awareness": "Problem Aware", "hypothesis": "...",
    "ads": [
      { "classification": "Imitation|Iteration|Ideation", "ad_format": "Video|Static|Promo",
        "copy": "...", "nota": "...", "imita_competidor": "(opcional)" }, ...
    ]
  }, ...
]

Reglas:
- 1 bache = 1 fila-cabecera (metadata D..I) + N filas de anuncio.
- En cada fila de anuncio: J Ad Type, K Ad Format, R copy; la "nota" (+ imita_competidor) va como COMENTARIO de celda en R.
- No destruye lo existente: arranca en start_row (si no se da, en la primera fila vacía por Status/BATCH#).
"""
import sys, json
from openpyxl import load_workbook
from openpyxl.comments import Comment

ATYPE = {"Imitation": "🎭 Imitation", "Iteration": "🔄 Iteration",
         "Ideation": "💡 Ideation", "Kalodata": "🦥 Kalodata"}
AFMT = {"Video": "🎬 Video", "Static": "🖼️ Static", "Promo": "🏷️ Promo"}
# columnas (1-indexed) de la convención Creative Roadmap
COL = {"status": 1, "batch": 2, "author": 3, "concept": 4, "angle": 5, "avatar": 6,
       "desire": 7, "awareness": 8, "hypothesis": 9, "adtype": 10, "adformat": 11, "copy": 18}


def first_empty_row(ws, start=3, maxscan=5000):
    r = start
    while r < maxscan:
        if not ws.cell(r, COL["status"]).value and not ws.cell(r, COL["batch"]).value \
           and not ws.cell(r, COL["copy"]).value:
            return r
        r += 1
    return r


def main():
    xlsx = sys.argv[1]
    batches = json.load(open(sys.argv[2], encoding="utf-8"))
    sheet = sys.argv[3] if len(sys.argv) > 3 else "Creative Roadmap"
    author = sys.argv[5] if len(sys.argv) > 5 else "AI"
    wb = load_workbook(xlsx)
    ws = wb[sheet]
    start = int(sys.argv[4]) if len(sys.argv) > 4 and str(sys.argv[4]).isdigit() else first_empty_row(ws)

    r = start
    for i, b in enumerate(batches, 1):
        ads = b.get("ads", [])
        for j, ad in enumerate(ads):
            if j == 0:
                ws.cell(r, COL["status"], "Ideando")
                ws.cell(r, COL["author"], author)
                ws.cell(r, COL["concept"], b.get("concept", ""))
                ws.cell(r, COL["angle"], b.get("angle", ""))
                ws.cell(r, COL["desire"], b.get("mass_desire", ""))
                ws.cell(r, COL["awareness"], b.get("awareness", ""))
                ws.cell(r, COL["hypothesis"], b.get("hypothesis", ""))
            ws.cell(r, COL["batch"], f"BATCH #{i}")
            ws.cell(r, COL["avatar"], b.get("avatar", ""))
            ws.cell(r, COL["adtype"], ATYPE.get(ad.get("classification", ad.get("ad_type", "Imitation")), "🎭 Imitation"))
            ws.cell(r, COL["adformat"], AFMT.get(ad.get("ad_format", "Video"), "🎬 Video"))
            cell = ws.cell(r, COL["copy"], ad.get("copy", ""))
            nota = ad.get("nota", "")
            if ad.get("imita_competidor"):
                nota = f"IMITA: {ad['imita_competidor']}\n\n{nota}"
            if nota:
                try:
                    cell.comment = Comment(nota[:2000], author)
                except Exception:
                    pass
            r += 1

    wb.save(xlsx)
    print(f"OK: {sum(len(b.get('ads', [])) for b in batches)} filas escritas en '{sheet}' (desde fila {start}).")


if __name__ == "__main__":
    main()
